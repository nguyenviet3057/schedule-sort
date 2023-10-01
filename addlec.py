from tkinter import ttk,messagebox,Tk
import openpyxl
import Guilec
import os

# GUI để thêm giáo viên 
class Addlec:
    def __init__(self,root,nguoigoi,classgoi,sheetsl):
        
        self.path = os.getcwd() + "\\alpha.xlsx"
        self.sheetsl = sheetsl
        self.classgoi = classgoi
        self.nguoigoi = nguoigoi
        self.root = root

        self.root.title('Thêm giảng viên')
        self.root.geometry('400x120+1500+50')

        self.frame = ttk.Frame(root)
        self.frame.pack()

        self.widgets_frame = ttk.LabelFrame(self.frame, text= "Giảng viên")
        self.widgets_frame.grid(row=0,column=0, padx=20, pady=10)

        self.name_entry = ttk.Entry(self.widgets_frame,font=("Helvetica", 20))
        self.name_entry.insert(0,"Tên")
        self.name_entry.bind("<FocusIn>", lambda e: self.name_entry.delete('0', 'end'))
        self.name_entry.grid(row=0,column=0,sticky='ew')

        button = ttk.Button(self.widgets_frame, text="Thêm", command=self.insert_col)
        button.grid(row=1, column=0, padx=5, pady=5, sticky="nsew")

        self.name_entry.bind("<Return>", self.perform_insert)
    
    def perform_insert(self, event=None):
        self.insert_col()

    def insert_col(self):
        name = self.name_entry.get()
        workbook = openpyxl.load_workbook(self.path)
        sheet = workbook[self.sheetsl]
        row = list(sheet.values)

        if len(row) == 0:
            sheet.cell(column = 1, row = 1,value = "Môn")
            sheet.cell(column = 2, row = 1,value = name)
        else:
            for i in row[0]:
                if i.lower() == name:
                    messagebox.showerror(title="Lỗi",message="Giảng viên đã tồn tại")
                    return
                else:
                     sheet.cell(column = len(row[0]) + 1, row = 1,value = name)               

        workbook.save(self.path)
        
        if self.classgoi.__name__ == "Guigiangvien" and (len(self.socot()) != 0 and len(self.sohang()) != 0):
            Guilec.Guigiangvien(self.nguoigoi,self.sheetsl,Guilec).close(self.nguoigoi)
            Guilec.Guigiangvien(Tk(),self.sheetsl,Guilec)

        workbook.close()
        self.root.destroy()
    
    def sohang(self):
        row = []
        workbook = openpyxl.load_workbook(self.path)
        sheet = workbook[self.sheetsl]
        data = list(sheet.values)
        if len(data) == 0:
            return row
        else:
            row = data[0]
            return row[1:]
    
    def socot(self):
        col = []
        workbook = openpyxl.load_workbook(self.path)
        sheet = workbook[self.sheetsl]
        data = list(sheet.values)
        for i in data[1:]:
            col.append(i[0])
        return col

# if __name__ == "__main__":
#     root = Tk()
#     Addlec(root,root,Guilec,"None")
#     root.mainloop()
