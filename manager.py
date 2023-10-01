from openpyxl import Workbook, load_workbook
import xlrd
import lichhoc
import readalpha as ra
import math

list_ = []
listtest = []

# Tìm cột STT (vị trí)
def colSTT(excel):
    for i in range(1,excel.max_row): # type: ignore
        for j in range(1,excel.max_column):
            val = excel.cell(row = i, column = j).value # type: ignore
            if val == "STT":
                return j

# Tìm cột tiêu đề (vị trí)
def rowheading(excel):
    for i in range(1,excel.max_row): # type: ignore
        for j in range(1,excel.max_column):
            val = excel.cell(row = i, column = j).value # type: ignore
            if val == "STT":
                return i
            
# Tìm lớp theo stt
def timlop(sche,stt):
    for i in sche:
        if i._stt == stt:
            return i

# Lấy số môn dạy của giảng viên ở môn đó
def somonday(sche,gv,tenmon):
    count = 0
    for i in sche:
        if i._course_name.strip().lower() == tenmon.strip().lower() and i._lec == gv:
            count = count + 1
    return count

# Xếp giảng viên (ưu tiên xếp hết)
def xepgiangvien(sche,sheetsl):
    listalpha = ra.listalpha(sheetsl)
    
    for tenmon in listalpha:
        index = 0
        somon = 0
        listlec = list(listalpha[tenmon].keys())
        listpoint = list(listalpha[tenmon].values())

        for lich in sche:
            if tenmon.lower().strip() == lich._course_name.lower().strip() and lich._lec == None and lich._course_name.strip().lower().count("đồ án") == 0:
                if len(listalpha[tenmon]) == 3:
                    # Tính tổng số môn mà gv được dạy
                    if index == 0:
                        maxx = math.floor(countmon(tenmon,sche)/3) + (countmon(tenmon,sche) % 3) + (int(listpoint[index]) - 4)
                    else:
                        maxx = math.floor(countmon(tenmon,sche)/3) + (int(listpoint[index]) - 4)

                    # Xét nếu mà có giảng viên ở lịch đó rồi không xếp nữa 
                    if lich._lec == None:
                        lich._lec = listlec[index]
                        somon = somon + 1

                        # Xếp gv cho lớp ghép
                        for i in sche:
                            if lich._lopghep == i._stt and i._lec == None:
                                i._lec = listlec[index]
                                somon = somon + 1
                                # Xếp gv cho lớp có 2 lịch
                                for j in sche:
                                    if j._course_name == i._course_name and j._class_name == i._class_name and j._lec == None:
                                        j._lec = listlec[index]
                                        somon = somon + 1

                        # Xếp gv cho lớp có 2 lịch dạy
                        for i in sche:
                            if lich._course_name == i._course_name and lich._class_name == i._class_name and i._lec == None:
                                i._lec = listlec[index]
                                somon = somon + 1

                        # Xếp gv cho lớp đồ án
                        for i in sche:
                            if i._course_name.strip().lower().count(lich._course_name.strip().lower()) > 0 and i._course_name.strip().lower().count("đồ án") != 0 and i._class_name == lich._class_name:
                                i._lec = listlec[index]

                    # Số môn gv dạy quá maxx thì chuyển xếp gv khác    
                    if somon >= maxx:
                        index = index + 1    
                        somon = 0

                if len(listalpha[tenmon]) == 2:
                    # Tính tổng số môn mà gv được dạy
                    if index == 0:
                        maxx = math.ceil(countmon(tenmon,sche)/2) + (int(listpoint[index]) - 6)
                    else:
                        maxx = math.floor(countmon(tenmon,sche)/2) + (int(listpoint[index]) - 6)
                    if maxx < 0:
                        maxx = 1

                    # Xét nếu mà có giảng viên ở lịch đó rồi không xếp nữa     
                    if lich._lec == None:
                        if index < len(listlec):
                            lich._lec = listlec[index]
                            somon = somon + 1
                            # Xếp gv cho lớp ghép
                            for i in sche:
                                if lich._lopghep == i._stt and i._lec == None:
                                    i._lec = listlec[index]
                                    somon = somon + 1
                                    # Xếp gv cho lớp có 2 lịch
                                    for j in sche:
                                            if j._course_name == i._course_name and j._class_name == i._class_name and j._lec == None:
                                                j._lec = listlec[index]
                                                somon = somon + 1
                                
                            # Xếp gv cho lớp có 2 lịch dạy
                            for i in sche:
                                if lich._course_name == i._course_name and lich._class_name == i._class_name and i._lec == None:
                                    i._lec = listlec[index]
                                    somon = somon + 1
                                
                            # Xếp gv cho những môn đồ án
                            for i in sche:
                                # Nếu có tên môn lại thêm chữ đồ án
                                if i._course_name.strip().lower().count(lich._course_name.strip().lower()) > 0 and i._course_name.strip().lower().count("đồ án") != 0 and i._class_name == lich._class_name and i._lec == None:
                                    i._lec = listlec[index]

                    # Số môn gv dạy quá maxx thì chuyển xếp gv khác  
                    if somon >= maxx:
                        index = index + 1
                        somon = 0
                    
                if len(listalpha[tenmon]) == 1:
                    lich._lec = listlec[0]
                    for i in sche:
                        if i._course_name.strip().lower().count(lich._course_name.strip().lower()) > 0 and i._class_name == lich._class_name:
                            i._lec = listlec[0]
    
    # Xếp nốt nhưng môn đồ án chưa được xếp
    for i in sche:
        if i._lec == None and i._course_name.strip().lower().count("đồ án") != 0:
            for j in sche:
                if i != j and i._class_name == j._class_name and i._course_name.strip().lower().count(j._course_name.strip().lower()) != 0 and j._course_name.strip().lower().count("đồ án") == 0:
                    i._lec = j._lec

    # for tenmon in listalpha:
    #     for lich in sche:
    #         if tenmon.lower() == lich._course_name.lower():
    #             print(tenmon,lich._lec)
    #     print("____")

# Xếp giảng viên (ưu tiên xếp theo alpha)
def xepgiangvien1(sche,sheetsl):
    listalpha = ra.listalpha(sheetsl)
    
    for tenmon in listalpha:
        listlec = list(listalpha[tenmon].keys())
        listpoint = list(listalpha[tenmon].values())

        for lich in sche:
            if tenmon.lower().strip() == lich._course_name.lower().strip() and lich._lec == None and lich._course_name.strip().lower().count("đồ án") == 0:
                if len(listalpha[tenmon]) == 3:

                    for vitri in range(len(listlec)):
                        # Tính tổng số môn mà gv được dạy
                        maxx = math.floor(countmon(tenmon,sche)/3) + (int(listpoint[vitri]) - 4)
                        if maxx <= 0:
                            maxx = 1

                        # Dự báo số môn sau khi xếp lịch xem có quá tổng số môn được dạy không thì mới xếp
                        dubao =  lich._2lich
                        if lich._lopghep != None:
                            dubao = dubao + timlop(sche,lich._lopghep)._2lich

                        if dubao + somonday(sche,listlec[vitri],tenmon) <= maxx:
                            # Xét nếu mà có giảng viên ở lịch đó rồi không xếp nữa 
                            if lich._lec == None:
                                lich._lec = listlec[vitri]                              

                                # Xếp gv cho lớp ghép
                                for i in sche:
                                    if lich._lopghep == i._stt and i._lec == None:
                                        i._lec = listlec[vitri]
                                        
                                        # Xếp gv cho lớp có 2 lịch
                                        for j in sche:
                                            if j._course_name == i._course_name and j._class_name == i._class_name and j._lec == None:
                                                j._lec = listlec[vitri]
                                                
                                # Xếp gv cho lớp có 2 lịch dạy
                                for i in sche:
                                    if lich._course_name == i._course_name and lich._class_name == i._class_name and i._lec == None:
                                        i._lec = listlec[vitri]
                                        
                                # Xếp gv cho lớp đồ án
                                for i in sche:
                                    if i._course_name.strip().lower().count(lich._course_name.strip().lower()) > 0 and i._course_name.strip().lower().count("đồ án") != 0 and i._class_name == lich._class_name:
                                        i._lec = listlec[vitri]
                                break

                if len(listalpha[tenmon]) == 2:

                    for vitri in range(len(listlec)):
                        # Tính tổng số môn mà gv được dạy
                        maxx = math.floor(countmon(tenmon,sche)/2) + (int(listpoint[vitri]) - 6)
                        if maxx <= 0:
                            maxx = 1

                        # Dự báo số môn sau khi xếp lịch xem có quá tổng số môn được dạy không thì mới xếp
                        dubao =  lich._2lich
                        if lich._lopghep != None:
                            dubao = dubao + timlop(sche,lich._lopghep)._2lich

                        if dubao + somonday(sche,listlec[vitri],tenmon) <= maxx:
                        # Xét nếu mà có giảng viên ở lịch đó rồi không xếp nữa     
                            if lich._lec == None:
                                lich._lec = listlec[vitri]

                                # Xếp gv cho lớp ghép
                                for i in sche:
                                    if lich._lopghep == i._stt and i._lec == None:
                                        i._lec = listlec[vitri]

                                        # Xếp gv cho lớp có 2 lịch
                                        for j in sche:
                                                if j._course_name == i._course_name and j._class_name == i._class_name and j._lec == None:
                                                    j._lec = listlec[vitri]
                                    
                                # Xếp gv cho lớp có 2 lịch dạy
                                for i in sche:
                                    if lich._course_name == i._course_name and lich._class_name == i._class_name and i._lec == None:
                                        i._lec = listlec[vitri]
                                    
                                # Xếp gv cho những môn đồ án
                                for i in sche:
                                    # Nếu có tên môn lại thêm chữ đồ án
                                    if i._course_name.strip().lower().count(lich._course_name.strip().lower()) > 0 and i._course_name.strip().lower().count("đồ án") != 0 and i._class_name == lich._class_name and i._lec == None:
                                        i._lec = listlec[vitri]
                                break
                        
                if len(listalpha[tenmon]) == 1:
                    lich._lec = listlec[0]
                    for i in sche:
                        if i._course_name.strip().lower().count(lich._course_name.strip().lower()) > 0 and i._class_name == lich._class_name:
                            i._lec = listlec[0]
    
    # Xếp nốt nhưng môn đồ án chưa được xếp
    for i in sche:
        if i._lec == None and i._course_name.strip().lower().count("đồ án") != 0:
            for j in sche:
                if i != j and i._class_name == j._class_name and i._course_name.strip().lower().count(j._course_name.strip().lower()) != 0 and j._course_name.strip().lower().count("đồ án") == 0:
                    i._lec = j._lec

    # for tenmon in listalpha:
    #     for lich in sche:
    #         if tenmon.lower() == lich._course_name.lower():
    #             print(tenmon,lich._lec)
    #     print("____")    


# Xếp lớp ghép
def checklopghep(sche):
    for i in sche:
        if i._lopghep == None:
            for j in sche:
                # Cùng mã học phần, cùng thứ, cùng tiết
                if j._lopghep == None:
                    if(i._course_code == j._course_code and i._day == j._day and i._session == j._session and i != j):
                        i._lopghep = j._stt
                        j._lopghep = i._stt
                        break

# Tìm những lớp mà giáo viên bị dạy trùng (loại trừ lớp ghép)
def checktrung(sche):
    for i in sche:
        for j in sche:
            # Cùng giáo viên, cùng thứ, cùng tiết, khác lớp ghép thì xét
            if i != j and i._lec == j._lec and i._day == j._day and i._session == j._session and i._stt != j._lopghep and i._lopghep != j._stt:
                if i._lec != None and j._lec != None:
                    same = ""
                    index = 0
                    for z in range(len("1234567890123456789012")):
                        if len(i._week[z]) < 22:
                            i._week = i._week + (22-len(i._week[z]))*" " 
                        if len(j._week[z]) < 22:
                            j._week = j._week + (22-len(j._week[z]))*" "     
                        if i._week[z] == j._week[z] and i._week[z] != " " and j._week[z] != " ":
                            # print(sche[i]._stt, sche[j]._stt,end="/")
                            index = j._stt
                            # same.append(j._week[z])
                            same = same + j._week[z]
                            # print(sche[i]._trung,sche[j]._trung,end="/")
                            # print(sche[i]._week[z])
                        else:
                            same = same + " "
                    if index != 0:
                        i._trung.append(index)        
                        i._tuantrung[j._stt] = same

# Tìm những lớp có 2 lịch ở 1 môn
def check2lich(sche):
    for i in sche:
        count = 1
        ds = []
        if i._2lich == 1:
            for j in sche:
                if i._course_code == j._course_code and i._class_name == j._class_name and j._2lich == 1 and i != j:
                    count = count + 1
                    ds.append(j)
                    i._2lich = count
            if len(ds) != 0:
                for i in ds:
                    i._2lich = count

# Tìm ngày bắt đầu năm học
def startday(sche):
    day = sche[0]._start
    for i in sche:
        if day > i._start:
            day = i._start
    return day

# Tìm ngày kết thúc năm học
def endday(sche):
    day = sche[0]._end
    for i in sche:
        if day < i._end:
            day = i._end
    return day

# Danh sách các môn học có tên là name (tạm thời không dùng đến nữa - thừa)
def countmon(name,sche):
    list_monhoc = []
    for i in sche:
        if i._course_name.lower() == name.lower():
            list_monhoc.append(i)
    return len(list_monhoc)

def canchecktrung(sche):
    for i in sche:
        if i != None:
            return True
    return False

# Đọc File .xlxs (Hàm: truyền đường dẫn trả về list)
def readfile(path):
    ds = []
    wb = load_workbook(path)
    sh = wb.active
    for i in range(1,sh.max_row): # type: ignore
        val = sh.cell(row = i, column = colSTT(sh)).value # type: ignore
        # Kiểm tra có phải dòng cần đọc không
        if isinstance(val, int) or val is not None and val.isdigit():
            for j in range(1, sh.max_column + 1):
                # Đọc theo tên cột (File excel phải đúng tên cột)
                match sh.cell(row = rowheading(sh), column = j).value:
                    case "STT":
                        stt = sh.cell(row = i, column = j).value
                    case "Mã học phần":
                        cousre_code = sh.cell(row = i, column = j).value   
                    case "Tên môn học":
                        course_name = sh.cell(row = i, column = j).value
                    case "Mã lớp học":
                        class_name = sh.cell(row = i, column = j).value
                    case "Lớp ghép":
                        class_com = sh.cell(row = i, column = j).value
                    case "Thứ":
                        day = sh.cell(row = i, column = j).value
                    case "Tiết":
                        seesion = sh.cell(row = i, column = j).value
                    case "Phòng học":
                        room = sh.cell(row = i, column = j).value
                    case "Số TC":
                        credit = sh.cell(row = i, column = j).value
                    case "Bắt đầu":
                        start = sh.cell(row = i, column = j).value
                    case "Kết thúc":
                        end = sh.cell(row = i, column = j).value
                    case "1234567890123456789012":
                        week = sh.cell(row = i, column = j).value
                    case "Giảng viên":
                        lec = sh.cell(row = i, column = j).value
            mh = lichhoc.monhoc(stt,cousre_code,course_name,class_name,class_com,day,seesion,room,credit,start,end,week)
            ds.append(mh)
    wb.close()
    return ds      

# Chuyển lưu dạng đối tượng về dạng list
def list_print(sche):
    ds = []
    for i in sche:
        x = [
            i._stt,
            i._course_code,
            i._course_name,
            i._class_name,
            i._day,
            i._session,
            i._week,
            i._lec,
            i._lopghep,
            i._trung,
        ]
        ds.append(x)
    return ds

def list_save(sche):
    ds = []
    for i in sche:
        x = [
            i._stt,
            i._course_code,
            i._course_name,
            i._class_name,
            i._class_com,
            i._day,
            i._session,
            i._room,
            i._credit,
            i._start,
            i._end,
            i._week,
            i._lec,
        ]
        ds.append(x)
    return ds

list_ = readfile('tkb.xlsx')

# checklopghep(list_)
# check2lich(list_)
# xepgiangvien(list_,"Sheet1")
# checktrung(list_)

# listtest = list_print(list_)

# print(listtest)



# Read File .xls
# workbook = xlrd.open_workbook('D:\\Code Python in VSC\\phan_cong.xls')
# sheet = workbook["Sheet2"]

# for i in range(sheet.nrows):
#     val = sheet.cell_value(i, colSTT1(sheet))
#     if isinstance(val, int) or (val is not None and str(val).isdigit()):
#         for j in range(1,sheet.ncols):
#             match sheet.cell_value(rowheading1(sheet), j):
#                 case "STT":
#                     stt = sheet.cell_value(i, j)
#                 case "Mã học phần":
#                     cousre_code = sheet.cell_value(i, j)   
#                 case "Tên môn học":
#                     course_name = sheet.cell_value(i, j)
#                 case "Mã lớp học":
#                     class_name = sheet.cell_value(i, j)
#                 case "Lớp ghép":
#                     class_com = sheet.cell_value(i, j)
#                 case "Thứ":
#                     day = sheet.cell_value(i, j)
#                 case "Tiết":
#                     seesion = sheet.cell_value(i, j)
#                 case "Phòng học":
#                     room = sheet.cell_value(i, j)
#                 case "Số TC":
#                     credit = sheet.cell_value(i, j)
#                 case "Bắt đầu":
#                     start = sheet.cell_value(i, j)
#                 case "Kết thúc":
#                     end = sheet.cell_value(i, j)
#                 case "1234567890123456789012":
#                     week = sheet.cell_value(i, j)
#         mh = lichhoc.monhoc(stt,cousre_code,course_name,class_name,class_com,day,seesion,room,credit,start,end,week)
#         list.append(mh)
