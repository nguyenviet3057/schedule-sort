import openpyxl
import os

path = os.getcwd() + "\\alpha.xlsx"

def ds(path,sheetsl):
    print(sheetsl)
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetsl]
    list_values = list(sheet.values)
    workbook.close()
    return list_values

# Danh sách các môn học trong alpha.xlsx
def listcourse():
    listmh = []
    for i in ds(path)[1:]:
        listmh.append(i[0].lower())
    return listmh

# Lưu giá trị môn nào nhưng giáo viên nào dạy và chỉ số alpha của cô đó
def listalpha(sheetsl):
    alpha = {}
    danhsach = ds(path,sheetsl)
    for row in range(1,len(danhsach)):
        lec = {}    
        for col in range(1,len(danhsach[0])):
            if danhsach[row][col] != None:
                lec[danhsach[0][col]] = danhsach[row][col]
        alpha[danhsach[row][0]] = lec
    return alpha

# Lịch dạy
def mhsche():
    import manager as mn
    x = []
    for i in mn.list:
        if i._course_name not in x:
            x.append(i._course_name)
    return x     

# Kiểm tra xem có thể sắp xếp được hay không( vì những môn trong lịch không có trong alpha.xlsx
# nên không biết chỉ số alpha bao nhiểu để sắp sếp)
def cansort():
    for i in mhsche():
        if i.lower() not in listcourse():
            print(i)
            return False
    return True    

def listsheet():
    if not os.path.exists(path):
        return []
    else:
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.sheetnames
        workbook.close()
        return sheet