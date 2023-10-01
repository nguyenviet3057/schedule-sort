from tkinter import ttk,messagebox,Tk
import openpyxl
import os
import readalpha as ra

# GUI để thêm năm học 
class Addnam:
    def __init__(self,root,nguoigoi,classgoi,sheetsl,updatecombobox,schecombo = None):
        
        self.path = os.getcwd() + "\\alpha.xlsx"
        self.nguoigoi = nguoigoi
        self.classgoi = classgoi
        self.sheetsl = sheetsl
        self.updatecombobox = updatecombobox
        self.schecombo = schecombo
        self.root = root

        self.root.title('Thêm năm học')
        self.root.geometry('400x120+1500+50')

        self.frame = ttk.Frame(root)
        self.frame.pack()

        self.widgets_frame = ttk.LabelFrame(self.frame, text= "Năm học")
        self.widgets_frame.grid(row=0,column=0, padx=20, pady=10)

        self.name_entry = ttk.Entry(self.widgets_frame,font=("Helvetica", 20))
        self.name_entry.insert(0,"Tên")
        self.name_entry.bind("<FocusIn>", lambda e: self.name_entry.delete('0', 'end'))
        self.name_entry.grid(row=0,column=0,sticky='ew')

        button = ttk.Button(self.widgets_frame, text="Thêm", command=self.insert_sheet)
        button.grid(row=1, column=0, padx=5, pady=5, sticky="nsew")

        self.name_entry.bind("<Return>", self.perform_insert)
    
    def perform_insert(self, event=None):
        self.insert_sheet()

    def insert_sheet(self):
        name = self.name_entry.get()

        if not os.path.exists(self.path):
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = name
            worksheet.cell(column = 1, row = 1,value = "Môn")
            workbook.save(self.path)
            workbook.close()

            self.root.destroy()
            self.updatecombobox()
            self.schecombo()
        else:
            for i in ra.listsheet():
                if i.lower() == name.lower():
                    messagebox.showerror(title="Lỗi",message="Năm học đã tồn tại")
                    return
                
            workbook = openpyxl.load_workbook(self.path)
            workbook.create_sheet(name)
            workbook.save(self.path)
            workbook.close()

            self.root.destroy()
            self.updatecombobox()
            self.schecombo()
            self.chepdulieu(name)

    def chepdulieu(self,name):
        answer = messagebox.askyesno(title="Gợi ý",message="Bạn có muốn sao chép năm học hiện tại sang năm học mới")
        data = ra.ds(self.path,self.sheetsl)
        workbook = openpyxl.load_workbook(self.path)
        worksheet = workbook[name]
        if (answer):
            for i in data:
                worksheet.append(i)
            workbook.save(self.path)
            workbook.close()
        else:
            worksheet.cell(column = 1, row = 1,value = "Môn")
            workbook.save(self.path)
            workbook.close()

# if __name__ == "__main__":
#     root = Tk()
#     Addnam(root,root,Addnam,"None",Addnam)
#     root.mainloop()