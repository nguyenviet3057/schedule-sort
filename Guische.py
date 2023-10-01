from tkinter import ttk,messagebox,filedialog,Menu,Tk
from openpyxl.styles import NamedStyle
import tkinter as tk
import openpyxl
import addlec
import addmon
import addnam
import manager as mn
import Guilec
import os
import readalpha as ra


#GUI hiện thị lịch
class Guische:
    def __init__(self,window):
        
        self.cansave = False
        self.path = os.getcwd() + "\\alpha.xlsx"
        self.list_copy = []

        self.window = window
        self.window.title('Lịch giảng dạy')
        self.window.geometry('+300+270')
    
        self.frame = ttk.Frame(self.window)
        self.frame.pack()

        style = ttk.Style(self.window)

        # self.window.tk.call("source", "forest-light.tcl")
        # self.window.tk.call("source", "forest-dark.tcl")
        # style.theme_use("forest-light")
        
        style.configure("Custom.TButton", font=("Helvetica", 13))

        self.widgets_frame = ttk.LabelFrame(self.frame, text= "Xử lý")
        self.widgets_frame.grid(row=0,column=0, padx=20, pady=10)

        self.name_entry = ttk.Entry(self.widgets_frame, font=("Helvetica", 14))
        self.name_entry.insert(0,"Alpha")
        # self.name_entry.bind("<FocusIn>", lambda e: self.name_entry.delete('0', 'end'))
        self.name_entry.grid(row=2,column=0,sticky='ew')

        self.button1 = ttk.Button(self.widgets_frame, text="Sửa", command=self.sua, style="Custom.TButton")
        self.button1.grid(row=3, column=0, padx=5, pady=5, sticky="nsew")

        self.status_combobox = ttk.Combobox(self.widgets_frame, values=ra.listsheet(), font=("Helvetica", 14))
        if len(ra.listsheet()) != 0:
            self.status_combobox.current(0)
        self.status_combobox.grid(row=4, column=0, padx=5, pady=5,  sticky="ew")

        self.button1 = ttk.Button(self.widgets_frame, text="Xếp giảng viên(Full)", command=self.xepgiangvien, style="Custom.TButton")
        self.button1.grid(row=5, column=0, padx=5, pady=5, sticky="nsew")

        self.button1 = ttk.Button(self.widgets_frame, text="Xếp giảng viên(Alpha)", command=self.xepgiangvien1, style="Custom.TButton")
        self.button1.grid(row=6, column=0, padx=5, pady=5, sticky="nsew")

        self.button1 = ttk.Button(self.widgets_frame, text="Kiểm tra lớp trùng", command=self.checkloptrung, style="Custom.TButton")
        self.button1.grid(row=7, column=0, padx=5, pady=5, sticky="nsew")

        self.button1 = ttk.Button(self.widgets_frame, text="Xuất File Excel", command=self.save, style="Custom.TButton")
        self.button1.grid(row=8, column=0, padx=5, pady=5, sticky="nsew")

        self.treeFrame = ttk.Frame(self.frame)
        self.treeFrame.grid(row=0, column=1, pady=10)
        self.treeScroll = ttk.Scrollbar(self.treeFrame)
        self.treeScroll.pack(side="right", fill="y")

        self.cols = ["STT","Mã học phần","Tên môn học", "Mã Lớp học phần", "Thứ","Tiết","Tuần","Giảng viên","Ghép","Trùng"]
        self.treeview = ttk.Treeview(self.treeFrame, show="headings",yscrollcommand=self.treeScroll.set, columns=self.cols, height=35)
        for i in self.cols:
            if i == "Tên môn học":
                self.treeview.column(i, width=330)
            elif i == "Tuần":
                self.treeview.column(i, width=160)
            else:
                self.treeview.column(i, width=100)

        self.treeview.pack()
        self.treeScroll.config(command=self.treeview.yview) 

        style = ttk.Style()
        style.configure("Treeview.Heading", font=("Helvetica", 12))
        self.treeview.tag_configure("my_font", font=("Helvetica", 12))

        def on_treeview_cell_select(event):
            region = self.treeview.identify_region(event.x, event.y)
            selected_item = self.treeview.selection()
            if region == "cell":
                selected_column = self.treeview.identify_column(event.x)  
                selected_value = self.treeview.item(selected_item[0], "values")[int(selected_column[1:]) - 1]
                self.name_entry.delete('0','end')
                self.name_entry.insert(0,selected_value)

        self.treeview.bind("<ButtonRelease-1>", on_treeview_cell_select)

        # self.load_data()

        self.menubar = Menu(window)
        self.window.config(menu = self.menubar)

        self.fileMenu = Menu(self.menubar, tearoff = 0)
        self.menubar.add_cascade(label = "File", menu = self.fileMenu)
        self.fileMenu.add_cascade(label = "Open",command= self.open)
        self.fileMenu.add_cascade(label = "Edit",command= self.edit)
        self.fileMenu.add_cascade(label = "Close",command= self.close)

        self.lecturerMenu = Menu(self.menubar, tearoff = 0)
        self.menubar.add_cascade(label = "Giảng viên", menu = self.lecturerMenu)
        self.lecturerMenu.add_cascade(label = "Thông Tin", command= self.thongtin)
        self.lecturerMenu.add_cascade(label = "Thêm Năm Học", command= self.addnamhoc)
        self.lecturerMenu.add_cascade(label = "Thêm Giảng Viên", command= self.addgiangvien)
        self.lecturerMenu.add_cascade(label = "Thêm Môn Học", command= self.addmonhoc)
    

    def close(self):
        self.window.quit()

    def open(self):
        self.file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx *.xls",)])
        if self.file_path:
            print("Thư mục được chọn:", self.file_path)
            self.treeview.delete(*self.treeview.get_children())

            self.list_ = mn.readfile(self.file_path)
            for i in self.list_:
                self.list_copy.append(i)

            mn.checklopghep(self.list_)
            mn.check2lich(self.list_)
            self.load_data(self.list_)
            self.cansave = True
        else:
            print("Không có thư mục nào được chọn.")

    def edit(self):
        self.treeview.delete(*self.treeview.get_children())

    def thongtin(self):
        selected_table = self.status_combobox.get()
        if not os.path.exists(self.path):
            messagebox.showwarning(title="Chú ý",message="Vui lòng thêm năm học")
            return
        if len(self.colfirst()) == 0 and len(self.rowfirst()) == 0:
            messagebox.showwarning(title="Chú ý",message="Không có thông tin giảng viên vui lòng thêm giảng viên và môn học")
        elif len(self.colfirst()) == 0 :
            messagebox.showwarning(title="Chú ý",message="Không có thông tin môn học vui lòng thêm môn học")
        elif len(self.rowfirst()) == 0 :
            messagebox.showwarning(title="Chú ý",message="Không có thông tin giảng viên vui lòng thêm giảng viên")
        else:
            Guilec.Guigiangvien(Tk(),selected_table,Guische,self.updatecombo)
            
    def updatecombo(self):
        self.status_combobox["value"] = ra.listsheet()

    def addnamhoc(self):
        selected_table = self.status_combobox.get()
        self.second_window = tk.Toplevel(self.window)
        addnam.Addnam(self.second_window,self.window,Guische,selected_table,self.updatecombo,self.updatecombo)

    def addgiangvien(self):
        # file_path = os.getcwd() + "\\alpha.xlsx"
        if os.path.exists(self.path):
            selected_table = self.status_combobox.get()
            root = Tk()
            addlec.Addlec(root,self.window,Guische,selected_table)
        else:
            messagebox.showwarning(title="Chú ý",message="Vui lòng thêm năm học")

    def addmonhoc(self):
        # file_path = os.getcwd() + "\\alpha.xlsx"
        if os.path.exists(self.path):
            selected_table = self.status_combobox.get()
            root = Tk()
            addmon.Addmonhoc(root,self.window,Guische,selected_table)
        else:
            messagebox.showwarning(title="Chú ý",message="Vui lòng thêm năm học")
    def sua(self):
        pass

    def creatfile():
        file_path = os.getcwd() + "\\alpha.xlsx"
        if not os.path.exists(file_path):
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "Sheet1"
            workbook.save(file_path)
            workbook.close()

    # Xếp giảng viên ưu tiên xếp hết
    def xepgiangvien(self):
        selected_table = self.status_combobox.get()
        print(selected_table)
        try:
            self.treeview.delete(*self.treeview.get_children())
            self.list_xep = mn.readfile(self.file_path)
            mn.checklopghep(self.list_xep)
            mn.check2lich(self.list_xep)
            mn.xepgiangvien(self.list_xep,selected_table)
            self.load_data(self.list_xep)
            self.list_ = self.list_xep
        except:
            messagebox.showwarning(title="Chú ý",message="Không có thông tin lịch dạy vui lòng thêm lịch dạy")
            return
        messagebox.showinfo(title="Thông báo",message="Xếp giảng viên xong")

    # Xếp giảng viên ưu tiên xếp theo alpha
    def xepgiangvien1(self):
        selected_table = self.status_combobox.get()
        try:
            self.treeview.delete(*self.treeview.get_children())
            self.list_xep = mn.readfile(self.file_path)
            mn.checklopghep(self.list_xep)
            mn.check2lich(self.list_xep)
            mn.xepgiangvien1(self.list_xep,selected_table)
            self.load_data(self.list_xep)
            self.list_ = self.list_xep
        except:
            messagebox.showwarning(title="Chú ý",message="Không có thông tin lịch dạy vui lòng thêm lịch dạy")
            return
        messagebox.showinfo(title="Thông báo",message="Xếp giảng viên xong")

    def checkloptrung(self):
        if mn.canchecktrung(self.list_):
            self.treeview.delete(*self.treeview.get_children())
            mn.checktrung(self.list_)
            self.load_data(self.list_)
        else:
           messagebox.showwarning(title="Chú ý",message="Không có thông tin giảng viên vui lòng thêm lịch dạy") 
           return
        messagebox.showinfo(title="Thông báo",message="Kiểm tra trùng lịch xong")
        
    def save(self):
        if self.cansave:
            file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
            if file_path:
                # Tạo một workbook và một worksheet
                workbook = openpyxl.Workbook()
                worksheet = workbook.active
                worksheet.title = "Sheet1"
                date_style = NamedStyle(name='date_style')
                date_style.number_format = 'DD-MM-YYYY'
                
                head = ["STT","Mã học phần","Tên môn học","Mã lớp học","Lớp ghép","Thứ","Tiết","Phòng học","Số TC","Bắt đầu","Kết thúc","1234567890123456789012","Giảng viên"]
                worksheet.append(head)
                list_data = mn.list_save(self.list_)
                for i in list_data:
                    worksheet.append(i)
                # Đổi style ngày
                column_J = worksheet['J']
                for cell in column_J:
                    cell.style = date_style
                
                column_K = worksheet['K']
                for cell in column_K:
                    cell.style = date_style
                # Lưu workbook
                workbook.save(file_path)
                workbook.close()
                messagebox.showinfo(title="Thông báo",message="Lưu File thành công")
        else:
            messagebox.showwarning(title="Chú ý",message="Không có thông tin lịch dạy vui lòng thêm lịch dạy")
    
    def load_data(self,data):
        colors = ["#E6F1D8", "white"]
        row_index = 0
        # data dạng đối tượng rồi mới chuyển thành dạng list
        list_data = mn.list_print(data)

        for col_name in self.cols:
            self.treeview.heading(col_name, text=col_name)

        chuoi = list_data[0][1] 
        swap = ""
        for value in list_data:
            if chuoi == value[1]:
                bg_color = colors[0]
            else:
                swap = colors[0]
                colors[0] = colors[1]
                colors[1] = swap

                bg_color = colors[0]
                chuoi = value[1]
            value_list = list(value)
            for i in range(len(value_list)):
                if value_list[i] == None:
                    value_list[i] = '-'
            self.treeview.insert('', tk.END, values=value_list, tags=("my_font",bg_color))

            row_index = row_index + 1

        for color in colors:
            self.treeview.tag_configure(color, background=color)

    def rowfirst(self):
        selected_table = self.status_combobox.get()
        row = []
        workbook = openpyxl.load_workbook(self.path)
        sheet = workbook[selected_table]
        data = list(sheet.values)
        if len(data) == 0:
            return row
        else:
            row = data[0]
            return row[1:]
    
    def colfirst(self):
        selected_table = self.status_combobox.get()
        col = []
        workbook = openpyxl.load_workbook(self.path)
        sheet = workbook[selected_table]
        data = list(sheet.values)
        for i in data[1:]:
            col.append(i[0])
        return col
    

if __name__ == "__main__":
    # Guische.creatfile()
    window = Tk()
    Guische(window)
    window.mainloop()