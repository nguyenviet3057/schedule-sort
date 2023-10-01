from tkinter import ttk,messagebox,Tk
import openpyxl
import Guilec
import os

# GUI đế thêm môn học 
class Addmonhoc:
    def __init__(self,root,nguoigoi,classgoi,sheetsl):
        
        self.path = os.getcwd() + "\\alpha.xlsx"
        self.sheetsl = sheetsl
        self.classgoi = classgoi
        self.nguoigoi = nguoigoi
        self.root = root
        self.root.geometry('400x120+1500+50')
        self.root.title('Thêm môn học')

        self.frame = ttk.Frame(self.root)
        self.frame.pack()

        self.widgets_frame = ttk.LabelFrame(self.frame, text= "Môn học")
        self.widgets_frame.grid(row=0,column=0, padx=20, pady=10)

        self.name_entry = ttk.Entry(self.widgets_frame,font=("Helvetica", 20))
        self.name_entry.insert(0,"Tên")
        self.name_entry.bind("<FocusIn>", lambda e: self.name_entry.delete('0', 'end'))
        self.name_entry.grid(row=0,column=0,sticky='ew')

        button = ttk.Button(self.widgets_frame, text="Thêm", command=self.insert_row)
        button.grid(row=1, column=0, padx=5, pady=5, sticky="nsew")

        self.name_entry.bind("<Return>", self.perform_insert)
    
    def perform_insert(self, event=None):
        self.insert_row()

    def insert_row(self):
        col = []
        name = self.name_entry.get()
        workbook = openpyxl.load_workbook(self.path)
        sheet = workbook[self.sheetsl]

        data = list(sheet.values)

        if len(data) == 0:
            row_values = ["Môn"]
            sheet.append(row_values)

        for i in data[1:]:
            col.append(i[0])
        for j in col:
            if j.lower() == name.lower():
                messagebox.showerror(title="Lỗi", message="Môn học đã tồn tại")
                return
            
        row_values = [name]
        sheet.append(row_values)
        workbook.save(self.path)

        if self.classgoi.__name__ == "Guigiangvien" and (len(self.socot()) != 0 and len(self.sohang()) != 0):
            Guilec.Guigiangvien(self.nguoigoi,self.sheetsl,Guilec).close(self.nguoigoi)
            Guilec.Guigiangvien(Tk(),self.sheetsl,Guilec)   

        workbook.close()
        self.root.destroy()
    
    def sohang(self):
        row = []
        workbook = openpyxl.load_workbook(self.path)
        sheet = workbook[self.sheetsl]
        data = list(sheet.values)
        if len(data) == 0:
            return row
        else:
            row = data[0]
            return row[1:]
    
    def socot(self):
        col = []
        workbook = openpyxl.load_workbook(self.path)
        sheet = workbook[self.sheetsl]
        data = list(sheet.values)
        for i in data[1:]:
            col.append(i[0])
        return col

# if __name__ == "__main__":
#     root = Tk()
#     Addmonhoc(root,root,"None")
#     root.mainloop()