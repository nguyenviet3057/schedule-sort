from tkinter import ttk,messagebox,Menu,Tk
import tkinter as tk
import openpyxl
import addlec
import addmon
import addnam
import readalpha as ra
import os

# GUI hiển thị chí số alpha giáo viên ứng với mỗi môn học
class Guigiangvien:

    def __init__(self,window,sheetsl,schegoi,schecombo = None):

        self.window = window
        self.sheetsl = sheetsl

        if schegoi.__name__ == "Guische":
            self.schecombo = schecombo

        self.path = os.getcwd() + "\\alpha.xlsx"

        self.window.title('Lịch giảng dạy')
        self.window.geometry('+0+0')

        self.frame = ttk.Frame(self.window)
        self.frame.pack()

        self.style = ttk.Style(self.window)

        # self.window.tk.call("source", "forest-light.tcl")
        # self.window.tk.call("source", "forest-dark.tcl")
        # self.style.theme_use("forest-light")

        # Khung thu 1

        self.style.configure("My.TLabelframe.Label", font=("Helvetica", 12))
        self.style.configure("My.TButton", font=("Helvetica", 12))

        self.widgets_frame = ttk.LabelFrame(self.frame, text= "Thông tin giảng viên", style="My.TLabelframe")
        self.widgets_frame.grid(row=0,column=0, padx=20, pady=10)

        self.status_combobox = ttk.Combobox(self.widgets_frame, values=self.rowfirst(self.sheetsl)[1:], font=("Helvetica", 14))
        self.status_combobox.current(0)
        self.status_combobox.grid(row=0, column=0, padx=5, pady=5,  sticky="ew")

        self.status_combobox1 = ttk.Combobox(self.widgets_frame, values=self.colfirst(self.sheetsl), font=("Helvetica", 14))
        self.status_combobox1.current(0)
        self.status_combobox1.grid(row=1, column=0, padx=5, pady=5,  sticky="ew")

        self.status_combobox2 = ttk.Combobox(self.widgets_frame, values=ra.listsheet(), font=("Helvetica", 14))
        # self.status_combobox2.current(0)
        self.status_combobox2.set(self.sheetsl)
        self.status_combobox2.grid(row=2, column=0, padx=5, pady=5,  sticky="ew")
        self.status_combobox2.bind("<<ComboboxSelected>>", self.switch_table)

        self.name_entry = ttk.Entry(self.widgets_frame, font=("Helvetica", 14))
        self.name_entry.insert(0,"Alpha")
        self.name_entry.bind("<FocusIn>", lambda e: self.name_entry.delete('0', 'end'))
        self.name_entry.grid(row=3,column=0,sticky='ew')

        self.button1 = ttk.Button(self.widgets_frame, text="Sửa", command=self.sua, style="My.TButton")
        self.button1.grid(row=4, column=0, padx=5, pady=5, sticky="nsew")

        self.separator = ttk.Separator(self.widgets_frame)
        self.separator.grid(row=5, column=0, padx=(20, 10), pady=10, sticky="ew")

        self.button1 = ttk.Button(self.widgets_frame, text="Kiểm tra", command=self.kiemtra, style="My.TButton")
        self.button1.grid(row=6, column=0, padx=5, pady=5, sticky="nsew")


        # Khung thu 2
        self.creattable(self.sheetsl,self.rowfirst(self.sheetsl))

        # Menu bar
        self.menubar = Menu(window)
        self.window.config(menu = self.menubar)

        self.lecturerMenu = Menu(self.menubar, tearoff = 0)
        self.menubar.add_cascade(label = "Giảng viên", menu = self.lecturerMenu)
        self.lecturerMenu.add_cascade(label = "Thông Tin", command= self.thongtin)
        self.lecturerMenu.add_cascade(label = "Thêm Năm Học", command= self.addnamhoc)
        self.lecturerMenu.add_cascade(label = "Thêm Giảng Viên", command= self.addgiangvien)
        self.lecturerMenu.add_cascade(label = "Thêm Môn Học", command= self.addmonhoc)

    def open():
        print("open file")

    def edit():
        print("edit")

    def close(self, window):
        self.window = window
        window.destroy()

    def thongtin(self):
        print(len(self.rowfirst(self.sheetsl)))

    def addnamhoc(self):
        selected_table = self.status_combobox2.get()
        self.second_window = tk.Toplevel(self.window)
        addnam.Addnam(self.second_window,self.window,Guigiangvien,selected_table,self.updatecombo,self.schecombo)
        
    def addgiangvien(self):
        selected_table = self.status_combobox2.get()
        root = Tk()
        addlec.Addlec(root,self.window,Guigiangvien,selected_table)

    def addmonhoc(self):
        selected_table = self.status_combobox2.get()
        root = Tk()
        addmon.Addmonhoc(root,self.window,Guigiangvien,selected_table)

    def colfirst(self,sheetsl):
        col = []
        workbook = openpyxl.load_workbook(self.path)
        sheet = workbook[sheetsl]
        data = list(sheet.values)
        for i in data[1:]:
            col.append(i[0])
        return col

    def rowfirst(self,sheetsl):
        workbook = openpyxl.load_workbook(self.path)
        sheet = workbook[sheetsl]
        row = list(sheet.values)
        return row[0]

    def load_data(self,sheetsl):
        colors = ["#E6F1D8", "white"]
        row_index = 0
        workbook = openpyxl.load_workbook(self.path)
        sheet = workbook[sheetsl]

        list_values = list(sheet.values)
        for col_name in list_values[0]:
            self.treeview.heading(col_name, text=col_name)

        for value in list_values[1:]:
            bg_color = colors[row_index % len(colors)]
            value_list = list(value)
            for i in range(len(value_list)):
                if value_list[i] == None:
                    value_list[i] = '-'
            self.treeview.insert('', tk.END, values=value_list, tags=("my_font",bg_color))
            row_index = row_index + 1
        
        for color in colors:
            self.treeview.tag_configure(color, background=color)

    # Sửa thông tin 
    def sua(self):  
        selected_table = self.status_combobox2.get()
        alpha = self.name_entry.get()
        giangvien = self.status_combobox.get()
        monhoc = self.status_combobox1.get()
        if alpha.isnumeric() or alpha == '':
            workbook = openpyxl.load_workbook(self.path)
            sheet = workbook[selected_table]
            sheet.cell(column = self.rowfirst(self.sheetsl).index(giangvien) + 1, row = self.colfirst(self.sheetsl).index(monhoc) + 2,value = alpha)
            workbook.save(self.path)
            self.treeview.delete(*self.treeview.get_children())
            self.load_data(self.sheetsl)
        else:
            messagebox.showerror(title = "Lỗi",message = "Vui lòng nhập số")

    # Kiểm tra xem tổng chỉ số alpha trong 1 môn có = 12 không
    def kiemtra(self):
        selected_table = self.status_combobox2.get()
        listalpha = ra.listalpha(selected_table)
        for tenmon in listalpha:
            tong = 0
            listpoint = list(listalpha[tenmon].values())
            for i in listpoint:
                tong = tong + int(i)
            if tong != 12:
                messagebox.showwarning(title = "Chú ý",message = tenmon + ": tổng hệ số alpha chưa = 12")
                break

    # Tạo bảng với dữ liệu và tiêu đề
    def creattable(self,table,heading):
        self.treeFrame = ttk.Frame(self.frame)
        self.treeFrame.grid(row=0, column=1, pady=10)
        self.treeScroll = ttk.Scrollbar(self.treeFrame)
        self.treeScroll.pack(side="right", fill="y")

        self.treeview = ttk.Treeview(self.treeFrame, show="headings",yscrollcommand=self.treeScroll.set, columns=heading, height=20)
        for i in heading:
            if i == "Môn":
                self.treeview.column(i, width=330)
            else:
                self.treeview.column(i, width=100)

        self.treeScroll.config(command=self.treeview.yview)
        # style = ttk.Style()
        self.style.configure("Treeview.Heading", font=("Helvetica", 12))
        self.treeview.tag_configure("my_font", font=("Helvetica", 12))
        self.load_data(table)
        self.treeview.pack()

        def select(event):
            indexrow = 0
            for i in self.treeview.selection():
                indexrow = self.treeview.index(i)
            self.status_combobox1.current(indexrow)

        def on_treeview_cell_select(event):
            region = self.treeview.identify_region(event.x, event.y)
            selected_item = self.treeview.selection()
            if region == "cell":
                selected_column = self.treeview.identify_column(event.x)  
                selected_column_id = int(selected_column[1:]) - 2
                selected_value = self.treeview.item(selected_item[0], "values")[int(selected_column[1:]) - 1]
                self.status_combobox.current(selected_column_id)
                if (selected_value.isnumeric()):
                    self.name_entry.delete('0','end')
                    self.name_entry.insert(0,selected_value)
                else:
                    self.name_entry.delete('0','end')  

        self.treeview.bind('<<TreeviewSelect>>',select)
        self.treeview.bind("<ButtonRelease-1>", on_treeview_cell_select)

    def switch_table(self,event):
    # Hiển thị TreeView tương ứng với giá trị Combobox
        selected_table = self.status_combobox2.get()
        listsheet = ra.listsheet()
        if len(self.socot()) == 0 and len(self.sohang()) == 0:
            messagebox.showwarning(title="Chú ý",message="Không có thông tin giảng viên vui lòng thêm giảng viên và môn học")
        elif len(self.socot()) == 0 :
            messagebox.showwarning(title="Chú ý",message="Không có thông tin môn học vui lòng thêm môn học")
        elif len(self.sohang()) == 0 :
            messagebox.showwarning(title="Chú ý",message="Không có thông tin giảng viên vui lòng thêm giảng viên")
        else:
            for i in listsheet:
                if selected_table == i:
                    self.sheetsl = selected_table
                    self.treeview.destroy()
                    self.creattable(selected_table,self.rowfirst(selected_table))
                    self.status_combobox["value"] = self.rowfirst(self.sheetsl)[1:]
                    self.status_combobox.current(0)
                    self.status_combobox1["value"] = self.colfirst(self.sheetsl)
                    self.status_combobox1.current(0)
    def sohang(self):
        selected_table = self.status_combobox2.get()
        row = []
        workbook = openpyxl.load_workbook(self.path)
        sheet = workbook[selected_table]
        data = list(sheet.values)
        if len(data) == 0:
            return row
        else:
            row = data[0]
            return row[1:]
    
    def socot(self):
        selected_table = self.status_combobox2.get()
        col = []
        workbook = openpyxl.load_workbook(self.path)
        sheet = workbook[selected_table]
        data = list(sheet.values)
        for i in data[1:]:
            col.append(i[0])
        return col
    
    def updatecombo(self):
        self.status_combobox2["value"] = ra.listsheet()         

# if __name__ == "__main__":
#     window = Tk()
#     Guigiangvien(window,"Sheet1",Guigiangvien)
#     window.mainloop()